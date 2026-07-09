#!/usr/bin/env python3
"""生成 rfq-to-quotation demo 用的虚构英文询盘 demo_inquiry.xlsx。

覆盖的动态列识别场景（见 examples/README.md）：
1. 标准列头 + 有数据列（COMMODITY DESCRIPTION / Qty/Kgs 等）—— 直接翻译保留
2. 标准列头 + 空数据列（CNY/Kgs）—— 识别为报价填写列，标黄
3. 无列头 + 有数据列 —— 数据为 C75S / 65Mn 等钢材牌号，应被推断为"钢材牌号"
4. 产品覆盖 DIN 125（平垫圈）/ DIN 7980（弹簧锁紧垫圈）/ DIN 471（轴用弹性挡圈），
   其中部分行只给规格号（如 M8 / D8 / 轴径20），需要靠 references/din_dimensions.md 补全尺寸
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# (COMMODITY DESCRIPTION, spec_code, 无列头钢材牌号列, Surface treatment, Qty/Kgs, CNY/Kgs(留空))
ROWS = [
    ("FLAT WASHER DIN125-1A", "M8",  "C75S", "ZP", 2500, None),
    ("FLAT WASHER DIN125-1A", "M10", "C75S", "ZP", 1800, None),
    ("FLAT WASHER DIN125-1A", "M12", "65Mn", "ZP", 1200, None),
    ("SPRING LOCK WASHER DIN7980", "D8",  "65Mn", "ZP",  900, None),
    ("SPRING LOCK WASHER DIN7980", "D10", "65Mn", "ZP",  700, None),
    ("SPRING LOCK WASHER HEAVY DIN7980", "D12", "65Mn", "ZP",  500, None),
    ("RETAINING RING FOR SHAFT DIN471", "20", "65Mn", "BZP", 400, None),
    ("RETAINING RING FOR SHAFT DIN471", "25", "65Mn", "BZP", 350, None),
    ("RETAINING RING FOR SHAFT DIN471", "30", "SS304", "",   300, None),
]

INQUIRY_NO = "RFQ-DEMO-20260709"


def build_inquiry(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiry"

    ws["A1"] = f"RFQ No.: {INQUIRY_NO}"
    ws["A1"].font = Font(bold=True, size=12)

    header_row = 3
    # 列 F（钢材牌号）故意不给列头，模拟客户 Excel 里"空列头但有数据"的常见情况
    headers = [
        "COMMODITY DESCRIPTION",
        "Spec",
        None,  # 无列头，数据为钢材牌号
        "Surface treatment",
        "Qty/Kgs",
        "CNY/Kgs",  # 报价列，客户留空
    ]
    for col, h in enumerate(headers, start=1):
        if h is not None:
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = Font(bold=True)

    row = header_row + 1
    for desc, spec, steel, surface, qty, price in ROWS:
        ws.cell(row=row, column=1, value=desc)
        ws.cell(row=row, column=2, value=spec)
        ws.cell(row=row, column=3, value=steel)
        ws.cell(row=row, column=4, value=surface)
        ws.cell(row=row, column=5, value=qty)
        ws.cell(row=row, column=6, value=price)  # 留空
        row += 1

    for col, width in zip("ABCDEF", [34, 8, 10, 16, 10, 12]):
        ws.column_dimensions[col].width = width

    wb.save(path)


if __name__ == "__main__":
    path = os.path.join(OUT_DIR, "demo_inquiry.xlsx")
    build_inquiry(path)
    print(f"SUCCESS: {path}")
