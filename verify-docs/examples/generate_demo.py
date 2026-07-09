#!/usr/bin/env python3
"""生成 verify-docs demo 用的虚构 CI / PL 贸易单据，故意埋 3 处不一致。

不一致清单（对答案见 examples/README.md）：
1. 行 3（DIN 933 螺栓 M10x50）数量：CI=500，PL=5000（数量不符）
2. 行 6（DIN 934 螺母 M10）净重：CI=45.2kg，PL=42.5kg（净重不符）
3. 合计金额：CI 打印的 TOTAL 比逐行小计之和少 100.00（合计金额不符）
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# 8 个行项目：品名、规格、数量(pcs)、单价(USD/千件)、净重(kg)、毛重(kg)
ITEMS = [
    ("Hex Bolt DIN 933",     "M8x40",  2000, 45.0, 60.0, 63.0),
    ("Hex Nut DIN 934",      "M8",     3000, 12.0, 21.0, 22.5),
    ("Hex Bolt DIN 933",     "M10x50", 500,  68.0, 38.5, 40.2),   # 数量不符行
    ("Flat Washer DIN 125",  "M8",     5000, 6.0,  18.0, 19.0),
    ("Spring Washer DIN 127","M8",     4000, 4.0,  9.6,  10.2),
    ("Hex Nut DIN 934",      "M10",    2500, 16.0, 45.2, 47.0),   # 净重不符行（PL会写42.5）
    ("Hex Bolt DIN 933",     "M12x60", 800,  95.0, 44.0, 46.5),
    ("Flat Washer DIN 125",  "M10",    3500, 9.0,  21.7, 22.9),
]

COMPANY_SELLER = "AcmeBolt Manufacturing Co., Ltd."
COMPANY_BUYER = "Global Fastener Trading Pty Ltd"
INVOICE_NO = "CI-DEMO-20260709"
PL_NO = "PL-DEMO-20260709"


def build_ci(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "CI"

    ws["A1"] = "COMMERCIAL INVOICE"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Seller:"
    ws["B3"] = COMPANY_SELLER
    ws["A4"] = "Buyer:"
    ws["B4"] = COMPANY_BUYER
    ws["A5"] = "Invoice No.:"
    ws["B5"] = INVOICE_NO
    ws["A6"] = "Date:"
    ws["B6"] = "2026-07-09"

    header_row = 8
    headers = ["No.", "Description", "Spec", "Qty (PCS)", "Unit Price (USD/1000PCS)", "Amount (USD)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True)

    total_amount = 0.0
    row = header_row + 1
    for idx, (name, spec, qty, price, nw, gw) in enumerate(ITEMS, start=1):
        amount = round(qty / 1000 * price, 2)
        total_amount += amount
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=spec)
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=6, value=amount)
        row += 1

    total_row = row
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    correct_total = round(total_amount, 2)
    # 故意埋差异 3：合计金额与逐行之和不一致（打字错位少输入 100）
    wrong_total = round(correct_total - 100.00, 2)
    ws.cell(row=total_row, column=6, value=wrong_total).font = Font(bold=True)

    ws.cell(row=total_row + 2, column=1, value="Amount in words:")
    ws.cell(row=total_row + 2, column=2, value=f"SAY US DOLLARS {_amount_words(correct_total)} ONLY")

    for col, width in zip("ABCDEF", [6, 26, 12, 12, 24, 14]):
        ws.column_dimensions[col].width = width

    wb.save(path)
    return correct_total, wrong_total


def build_pl(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "PL"

    ws["A1"] = "PACKING LIST"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    ws["A3"] = "Seller:"
    ws["B3"] = COMPANY_SELLER
    ws["A4"] = "Buyer:"
    ws["B4"] = COMPANY_BUYER
    ws["A5"] = "Packing List No.:"
    ws["B5"] = PL_NO
    ws["A6"] = "Date:"
    ws["B6"] = "2026-07-09"

    header_row = 8
    headers = ["No.", "Description", "Spec", "Qty (PCS)", "N.W. (KG)", "G.W. (KG)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True)

    total_nw = 0.0
    total_gw = 0.0
    row = header_row + 1
    for idx, (name, spec, qty, price, nw, gw) in enumerate(ITEMS, start=1):
        pl_qty = qty
        pl_nw = nw
        if idx == 3:
            # 故意埋差异 1：数量不符（CI=500 -> PL=5000）
            pl_qty = 5000
        if idx == 6:
            # 故意埋差异 2：净重不符（CI=45.2 -> PL=42.5）
            pl_nw = 42.5
        total_nw += pl_nw
        total_gw += gw
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=spec)
        ws.cell(row=row, column=4, value=pl_qty)
        ws.cell(row=row, column=5, value=pl_nw)
        ws.cell(row=row, column=6, value=gw)
        row += 1

    total_row = row
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=round(total_nw, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(total_gw, 2)).font = Font(bold=True)

    for col, width in zip("ABCDEF", [6, 26, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = width

    wb.save(path)


def _amount_words(amount):
    """极简大写转换，仅供 demo 展示用，不追求财务级严谨。"""
    dollars = int(amount)
    cents = round((amount - dollars) * 100)
    return f"{dollars:,} AND CENTS {cents:02d}"


if __name__ == "__main__":
    ci_path = os.path.join(OUT_DIR, "demo_ci.xlsx")
    pl_path = os.path.join(OUT_DIR, "demo_pl.xlsx")
    correct_total, wrong_total = build_ci(ci_path)
    build_pl(pl_path)
    print(f"SUCCESS: {ci_path}")
    print(f"SUCCESS: {pl_path}")
    print(f"CI correct row-sum total = {correct_total}, CI printed (wrong) total = {wrong_total}")
