#!/usr/bin/env python3
"""
Market Intel 报告生成器 — 从线索 JSON 生成多 Sheet Excel。

设计要点：
- 直接读 JSON 文件（--leads leads.jsonl 或 --data '<json>'），避免脆弱的单行字符串转义
- 结构化写入 openpyxl，条件格式按 tier 着色
- 内置 round-trip 自验（生成后 openpyxl 重开核对行数/tier分布）

用法：
  python3 build_report.py --leads leads.jsonl --meta meta.json --output report.xlsx
  python3 build_report.py --data '<json数组>' --meta-data '<json>' --output report.xlsx

leads 结构见 references/lead-schema.md。meta 结构见 assets/report_template_notes.md。
成功打印 "SUCCESS: <路径>"，自验失败打印 "VERIFY_FAIL: <原因>" 并退出码 1。
"""
import argparse, json, re, sys
from datetime import datetime, timezone
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 评分权重（须与 references/scoring-rubric.md 保持同步；改一处务必改另一处） ──
W = {"email": 0.40, "cross": 0.25, "decision_maker": 0.20, "scale": 0.15}
PENALTY_CAP = 0.2
PENALTY_PER_FLAG = 0.1
A_SCORE_MIN = 0.75
B_SCORE_MIN = 0.5

PURCHASING_PREFIXES = ("purchasing", "procurement", "buying", "supply", "sourcing")
GENERIC_PREFIXES = ("sales", "info", "office")

# ── 样式 ──
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TIER_FILL = {"A": PatternFill("solid", fgColor="C6EFCE"),   # 绿
             "B": PatternFill("solid", fgColor="FFEB9C"),   # 黄
             "C": PatternFill("solid", fgColor="FFC7CE")}   # 红
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

DISCLAIMER = ("本报告基于公开网页信息，不含海关采购记录；联系方式可信度见各行 email_level 分级"
              "（E1官网具名/部门·可直接发信，E2官网generic·通用模板，E3推断未验证，E4无邮箱）。")


def _hdr(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"


def _best_email(dms):
    """取决策人里最佳邮箱 + 级别"""
    order = {"E1": 0, "E2": 1, "E3": 2, "E4": 3}
    cands = [(order.get(d.get("email_level", "E4"), 3), d) for d in dms if d.get("email")]
    if not cands:
        return "", "E4"
    cands.sort(key=lambda x: x[0])
    d = cands[0][1]
    return d.get("email", ""), d.get("email_level", "E4")


def _dm_summary(dms):
    parts = []
    for d in dms:
        nm = d.get("name") or d.get("role") or "generic"
        em = d.get("email", "")
        lv = d.get("email_level", "")
        parts.append(f"{nm}({d.get('role','')}) {em} [{lv}]".strip())
    return "\n".join(parts)


def _check_lead(obj):
    """校验单条线索的最小结构；返回问题描述，None 表示通过"""
    if not isinstance(obj, dict):
        return "不是 JSON 对象"
    if "company" not in obj:
        return "缺 company 字段"
    cr = obj.get("credibility")
    if not isinstance(cr, dict) or cr.get("tier") not in ("A", "B", "C"):
        return "credibility.tier 缺失或不是 A/B/C"
    if not isinstance(cr.get("score"), (int, float)):
        return "credibility.score 缺失或非数值"
    return None


def _domain(url_or_email, is_email=False):
    """从 URL 或 email 提取归一化域名（小写、剥离 www.）"""
    if not url_or_email:
        return ""
    if is_email:
        if "@" not in url_or_email:
            return ""
        host = url_or_email.rsplit("@", 1)[-1].strip().lower()
    else:
        u = url_or_email.strip()
        if "://" not in u:
            u = "http://" + u
        host = (urlparse(u).netloc or "").lower()
        host = host.split(":")[0]  # 去端口
    if host.startswith("www."):
        host = host[4:]
    return host


def _local_part(email):
    return email.split("@", 1)[0].strip().lower() if email and "@" in email else ""


def _verify_email_level(dm, homepage):
    """代码复算单个决策人邮箱的 email_level。返回 (level, critical_flag_or_None)。

    复算方向是**只收紧、不放宽**：E1/E2 与"邮箱在页面上明确出现"这一事实绑定，
    而"是否在页面出现"是 S2/S3 阶段人工判断的结果（体现在自报 email_level 里，
    E3 明确代表"推断·未在任何页面出现"），代码侧拿不到原始页面重新判断。
    所以：
    - 域名不匹配（剥 www. 后比较）→ 无论自报什么，强制降到 E3 + critical 红旗
      （这个方向代码能独立验证：域名比对是纯字符串运算，不依赖"是否曾出现在页面"这类人工判断）。
    - 域名匹配且自报已是 E1/E2 → 复核是否满足职能白名单/具名条件，不满足则降级
      （防止"随手挂个 E1 却连白名单前缀都不占"的虚报）。
    - 域名匹配但自报是 E3/E4 → 尊重原判（代码不能把"推断未验证"升级为"已验证"，
      那需要页面证据，只有人工/子agent持有）。
    - mx_ok=false（若子agent提供该字段）→ 一律降 E4。
    """
    email = (dm.get("email") or "").strip()
    if not email or "@" not in email:
        return "E4", None

    home_domain = _domain(homepage)
    email_domain = _domain(email, is_email=True)
    domain_matches = bool(home_domain) and email_domain == home_domain

    if not domain_matches:
        return "E3", "邮箱域名与官网域名不符（可能是代理/钓鱼/信息不一致）"

    reported = dm.get("email_level")
    local = _local_part(email)
    named = bool(dm.get("name"))
    qualifies_e1 = named or any(local.startswith(p) for p in PURCHASING_PREFIXES)
    qualifies_e2 = any(local.startswith(p) for p in GENERIC_PREFIXES)

    if reported in ("E3", "E4"):
        # 只收紧不放宽：代码没有页面证据把"未验证"升级为"已验证"
        level = reported
    elif reported == "E1":
        level = "E1" if qualifies_e1 else ("E2" if qualifies_e2 else "E2")
    else:
        # 自报 E2 或缺失自报值：域名匹配后按职能白名单归类
        level = "E1" if qualifies_e1 else "E2"

    if dm.get("mx_ok") is False:
        level = "E4"

    return level, None


def _email_component(level):
    return {"E1": 1.0, "E2": 0.7, "E3": 0.3, "E4": 0.0}.get(level, 0.0)


def _decision_maker_component(dms):
    """具名决策人=1.0 / 部门角色=0.6 / 仅generic=0.3 / 无=0.0"""
    has_named = any(d.get("name") for d in dms)
    if has_named:
        return 1.0
    has_dept = any(
        (not d.get("name")) and d.get("role") and d.get("role") != "generic" and d.get("email")
        for d in dms
    )
    if has_dept:
        return 0.6
    has_generic = any(d.get("email") for d in dms)
    if has_generic:
        return 0.3
    return 0.0


def _clamp(x, lo=0.0, hi=1.0):
    try:
        x = float(x)
    except (TypeError, ValueError):
        return lo
    return max(lo, min(hi, x))


def _recompute_credibility(lead, warnings):
    """代码复算 credibility：subscores 存在则加权复算+交叉核对；
    缺失则直接从原始特征复算（兼容旧数据）。返回是否发生了纠正（bool）。
    warnings: list，追加人类可读 WARN 文案（不含 print，调用方负责输出）。
    """
    lead_id = lead.get("lead_id", "<无id>")
    cr = lead.setdefault("credibility", {})
    dms = lead.get("decision_makers", [])
    homepage = lead.get("company", {}).get("homepage", "")

    # 1) 逐个决策人邮箱强制复算 email_level + 收集 critical 红旗
    critical_flags = list(cr.get("red_flags_critical", []) or [])
    for dm in dms:
        if not dm.get("email"):
            continue
        recomputed_level, critical_reason = _verify_email_level(dm, homepage)
        reported_level = dm.get("email_level")
        if reported_level != recomputed_level:
            warnings.append(
                f"{lead_id}: 决策人邮箱 {dm.get('email')} 自报级别 {reported_level} "
                f"!= 复算级别 {recomputed_level}，已覆盖"
            )
        dm["email_level"] = recomputed_level
        if critical_reason and critical_reason not in critical_flags:
            critical_flags.append(critical_reason)

    # 2) 取该线索最佳邮箱级别（复算后）
    order = {"E1": 0, "E2": 1, "E3": 2, "E4": 3}
    levels = [dm.get("email_level", "E4") for dm in dms if dm.get("email")]
    best_level = min(levels, key=lambda l: order.get(l, 3)) if levels else "E4"

    cross_count = lead.get("lead_source", {}).get("cross_source_count", 0) or 0
    minor_flags = cr.get("red_flags", []) or []
    penalty_from_features = min(len(minor_flags) * PENALTY_PER_FLAG, PENALTY_CAP)

    email_c = _email_component(best_level)
    cross_c = min(cross_count, 3) / 3
    dm_c = _decision_maker_component(dms)

    subscores = cr.get("subscores")
    if isinstance(subscores, dict):
        # subscores 存在：clamp 后按权重复算；并交叉核对与特征是否一致
        s_email = _clamp(subscores.get("email", email_c))
        s_cross = _clamp(subscores.get("cross", cross_c))
        s_dm = _clamp(subscores.get("decision_maker", dm_c))
        s_scale = _clamp(subscores.get("scale", 0.0))
        s_penalty = _clamp(subscores.get("penalty", penalty_from_features), 0.0, PENALTY_CAP)

        if abs(s_email - email_c) > 0.02:
            warnings.append(
                f"{lead_id}: subscores.email={s_email} 与复算邮箱子分 {email_c:.2f}（基于 {best_level}）不符，以特征为准"
            )
            s_email = email_c
        if abs(s_cross - cross_c) > 0.02:
            warnings.append(
                f"{lead_id}: subscores.cross={s_cross} 与复算交叉子分 {cross_c:.2f}（cross_source_count={cross_count}）不符，以特征为准"
            )
            s_cross = cross_c
        if abs(s_penalty - penalty_from_features) > 0.02:
            warnings.append(
                f"{lead_id}: subscores.penalty={s_penalty} 与 red_flags 数量算出的 {penalty_from_features:.2f} 不符，以特征为准"
            )
            s_penalty = penalty_from_features

        score = W["email"] * s_email + W["cross"] * s_cross + W["decision_maker"] * s_dm + W["scale"] * s_scale - s_penalty
        cr["subscores"] = {
            "email": round(s_email, 4), "cross": round(s_cross, 4),
            "decision_maker": round(s_dm, 4), "scale": round(s_scale, 4),
            "penalty": round(s_penalty, 4),
        }
    else:
        # subscores 缺失（旧数据兼容）：直接从特征复算；scale 无独立报值可复算，
        # 不臆造 —— 取 0（保守：宁可低估，也不凭空拔高）
        s_scale = 0.0
        score = W["email"] * email_c + W["cross"] * cross_c + W["decision_maker"] * dm_c + W["scale"] * s_scale - penalty_from_features
        cr["subscores"] = {
            "email": round(email_c, 4), "cross": round(cross_c, 4),
            "decision_maker": round(dm_c, 4), "scale": round(s_scale, 4),
            "penalty": round(penalty_from_features, 4),
        }

    score = _clamp(score)

    # 3) tier 复算（严格对齐 references/scoring-rubric.md 的判据，单一路径无覆盖）
    has_critical = len(critical_flags) > 0
    is_c = (score < B_SCORE_MIN) or (cross_count <= 1) or (best_level == "E4")

    if (score >= A_SCORE_MIN and best_level in ("E1", "E2")
            and cross_count >= 2 and not has_critical):
        tier = "A"
    elif is_c:
        tier = "C"
    else:
        # 未达 A 门槛、也未落入 C 判据（即 score>=0.5 且 cross>=2 且 best!=E4）→ B
        # 含：critical 红旗封顶 B；best=E3 且分数达标；分数在 [0.5,0.75) 区间
        tier = "B"

    reported_score = lead.get("credibility", {}).get("score")
    reported_tier = lead.get("credibility", {}).get("tier")
    corrected = False
    if reported_score is None or abs(_clamp(reported_score) - score) > 0.02 or reported_tier != tier:
        warnings.append(
            f"{lead_id}: 自报 score={reported_score}/tier={reported_tier} != 复算 score={round(score,2)}/tier={tier}，已覆盖"
        )
        corrected = True

    cr["score"] = round(score, 2)
    cr["tier"] = tier
    cr["email_best_level"] = best_level
    cr["red_flags_critical"] = critical_flags
    cr["red_flags"] = minor_flags

    if tier == "C" and not cr.get("next_verification_step"):
        cr["next_verification_step"] = "需人工补充验证步骤"
        warnings.append(f"{lead_id}: C 级缺 next_verification_step，已自动填占位文案")
        corrected = True

    return corrected


def build(leads, meta, output):
    wb = Workbook()

    # ── Sheet 1: 线索总览 ──
    ws = wb.active
    ws.title = "线索总览"
    cols = ["公司名", "国家", "tier", "score", "最佳邮箱", "邮箱级别",
            "决策人", "产品重合", "出单主体", "交叉源数", "推荐渠道", "切入策略"]
    _hdr(ws, cols)
    leads_sorted = sorted(leads, key=lambda x: ("ABC".index(x["credibility"]["tier"]),
                                                 -x["credibility"]["score"]))
    for r, ld in enumerate(leads_sorted, 2):
        co, cr, bd = ld["company"], ld["credibility"], ld.get("bd_strategy", {})
        email, lvl = _best_email(ld.get("decision_makers", []))
        row = [co.get("name_en") or co.get("name_local", ""), co.get("country", ""),
               cr["tier"], round(cr["score"], 2), email, lvl,
               _dm_summary(ld.get("decision_makers", [])),
               ", ".join(co.get("product_line_match", [])), co.get("target_entity", ""),
               ld.get("lead_source", {}).get("cross_source_count", 0),
               bd.get("recommended_channel", ""), bd.get("entry_angle", "")]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border, cell.alignment = BORDER, WRAP
        ws.cell(r, 3).fill = TIER_FILL.get(cr["tier"], PatternFill())
    _autosize(ws, {1: 26, 7: 34, 12: 40})

    # ── Sheet 2: A级·可直接发信 ──
    ws2 = wb.create_sheet("A级·可直接发信")
    a_cols = ["公司名", "国家", "官网", "决策人", "最佳邮箱", "产品重合",
              "切入角度", "模板提示", "母语渠道情报"]
    _hdr(ws2, a_cols)
    a_leads = [l for l in leads_sorted if l["credibility"]["tier"] == "A"]
    for r, ld in enumerate(a_leads, 2):
        co, bd = ld["company"], ld.get("bd_strategy", {})
        email, _ = _best_email(ld.get("decision_makers", []))
        row = [co.get("name_en") or co.get("name_local", ""), co.get("country", ""),
               co.get("homepage", ""), _dm_summary(ld.get("decision_makers", [])), email,
               ", ".join(co.get("product_line_match", [])), bd.get("entry_angle", ""),
               bd.get("template_hint", ""), ld.get("native_channel_intel", "")]
        for c, v in enumerate(row, 1):
            cell = ws2.cell(r, c, v); cell.border, cell.alignment = BORDER, WRAP
    _autosize(ws2, {1: 26, 3: 28, 4: 34, 7: 34, 8: 30, 9: 30})

    # ── Sheet 3: B/C级·待验证 ──
    ws3 = wb.create_sheet("B-C级·待验证")
    bc_cols = ["公司名", "国家", "tier", "score", "邮箱(级别)", "疑点", "下一步验证"]
    _hdr(ws3, bc_cols)
    bc_leads = [l for l in leads_sorted if l["credibility"]["tier"] in ("B", "C")]
    for r, ld in enumerate(bc_leads, 2):
        co, cr = ld["company"], ld["credibility"]
        email, lvl = _best_email(ld.get("decision_makers", []))
        row = [co.get("name_en") or co.get("name_local", ""), co.get("country", ""),
               cr["tier"], round(cr["score"], 2), f"{email} [{lvl}]",
               "; ".join(cr.get("red_flags", [])), cr.get("next_verification_step", "") or ""]
        for c, v in enumerate(row, 1):
            cell = ws3.cell(r, c, v); cell.border, cell.alignment = BORDER, WRAP
        ws3.cell(r, 3).fill = TIER_FILL.get(cr["tier"], PatternFill())
    _autosize(ws3, {1: 26, 6: 34, 7: 40})

    # ── Sheet 4: 母语渠道情报 ──
    ws4 = wb.create_sheet("母语渠道情报")
    _hdr(ws4, ["公司名", "国家", "母语渠道情报"])
    for r, ld in enumerate([l for l in leads_sorted if l.get("native_channel_intel")], 2):
        co = ld["company"]
        for c, v in enumerate([co.get("name_en") or co.get("name_local", ""),
                               co.get("country", ""), ld.get("native_channel_intel", "")], 1):
            cell = ws4.cell(r, c, v); cell.border, cell.alignment = BORDER, WRAP
    _autosize(ws4, {1: 26, 3: 60})

    # ── Sheet 5: 运行元数据 ──
    ws5 = wb.create_sheet("运行元数据")
    tiers = {"A": 0, "B": 0, "C": 0}
    for l in leads:
        tiers[l["credibility"]["tier"]] += 1
    rows = [
        ["报告生成时间", meta.get("generated_at", "")],
        ["关键词", ", ".join(meta.get("keywords", []))],
        ["HS编码", ", ".join(meta.get("hs_codes", []))],
        ["目标市场", ", ".join(meta.get("markets", []))],
        ["", ""],
        ["漏斗 · 发现候选", meta.get("funnel", {}).get("discovered", "")],
        ["漏斗 · 提取存活", meta.get("funnel", {}).get("extracted", "")],
        ["漏斗 · 背调完成", meta.get("funnel", {}).get("verified", "")],
        ["分级 · A(可直接发信)", tiers["A"]],
        ["分级 · B(待补验证)", tiers["B"]],
        ["分级 · C(仅供参考)", tiers["C"]],
        ["", ""],
        ["firecrawl credits 消耗", meta.get("cost", {}).get("firecrawl_credits", "")],
        ["数据源", ", ".join(meta.get("sources_used", []))],
        ["", ""],
        ["免责声明", DISCLAIMER],
    ]
    if meta.get("corrected_count"):
        rows.insert(-2, ["评分被代码纠正", meta["corrected_count"]])
    if meta.get("skipped_lines"):
        rows.insert(-2, ["跳过的坏行", meta["skipped_lines"]])
    for r, (k, v) in enumerate(rows, 1):
        kc = ws5.cell(r, 1, k); kc.font = Font(bold=True); kc.alignment = WRAP
        vc = ws5.cell(r, 2, v); vc.alignment = WRAP
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 70

    wb.save(output)
    return {"total": len(leads), "tiers": tiers, "a_count": len(a_leads),
            "bc_count": len(bc_leads)}


def _autosize(ws, overrides=None):
    overrides = overrides or {}
    for c in range(1, ws.max_column + 1):
        if c in overrides:
            ws.column_dimensions[get_column_letter(c)].width = overrides[c]
        else:
            ws.column_dimensions[get_column_letter(c)].width = 14


def verify(output, expect):
    """Round-trip 自验：重开核对行数和 tier 分布"""
    wb = load_workbook(output)
    problems = []
    # 总览行数 == 线索数
    n_overview = wb["线索总览"].max_row - 1
    if n_overview != expect["total"]:
        problems.append(f"总览行数 {n_overview} != 线索数 {expect['total']}")
    # A级 sheet 行数 == A级数
    n_a = wb["A级·可直接发信"].max_row - 1
    if n_a != expect["a_count"]:
        problems.append(f"A级sheet行数 {n_a} != A级数 {expect['a_count']}")
    # B/C sheet 行数
    n_bc = wb["B-C级·待验证"].max_row - 1
    if n_bc != expect["bc_count"]:
        problems.append(f"B-C级sheet行数 {n_bc} != B-C级数 {expect['bc_count']}")
    # 免责声明存在
    meta_vals = [wb["运行元数据"].cell(r, 2).value for r in range(1, wb["运行元数据"].max_row + 1)]
    if not any(v and "不含海关采购记录" in str(v) for v in meta_vals):
        problems.append("元数据缺免责声明")
    return problems


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--leads", help="JSONL 文件路径（每行一条线索）")
    ap.add_argument("--data", help="JSON 数组字符串（备选）")
    ap.add_argument("--meta", help="meta JSON 文件路径")
    ap.add_argument("--meta-data", help="meta JSON 字符串（备选）")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    skipped = 0
    if args.leads:
        leads = []
        with open(args.leads, encoding="utf-8") as f:
            for i, line in enumerate(f, 1):
                if not line.strip():
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError as e:
                    print(f"WARN: 第{i}行 JSON 解析失败，已跳过: {e}", file=sys.stderr)
                    skipped += 1
                    continue
                problem = _check_lead(obj)
                if problem:
                    print(f"WARN: 第{i}行{problem}，已跳过", file=sys.stderr)
                    skipped += 1
                    continue
                leads.append(obj)
    elif args.data:
        try:
            parsed = json.loads(args.data)
        except json.JSONDecodeError as e:
            print(f"VERIFY_FAIL: --data 不是合法 JSON: {e}", file=sys.stderr); sys.exit(1)
        if not isinstance(parsed, list):
            print("VERIFY_FAIL: --data 必须是 JSON 数组", file=sys.stderr); sys.exit(1)
        leads = []
        for i, obj in enumerate(parsed, 1):
            problem = _check_lead(obj)
            if problem:
                print(f"WARN: 第{i}条{problem}，已跳过", file=sys.stderr)
                skipped += 1
                continue
            leads.append(obj)
    else:
        print("VERIFY_FAIL: 需 --leads 或 --data", file=sys.stderr); sys.exit(1)

    if args.meta:
        with open(args.meta, encoding="utf-8") as f:
            meta = json.load(f)
    elif args.meta_data:
        meta = json.loads(args.meta_data)
    else:
        meta = {}
    meta.setdefault("generated_at", datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M"))

    if not leads:
        msg = f"leads 无有效数据（共跳过 {skipped} 行）" if skipped else "线索为空"
        print(f"VERIFY_FAIL: {msg}", file=sys.stderr); sys.exit(1)
    if skipped:
        meta["skipped_lines"] = skipped

    # 代码复算强制：逐条以 references/scoring-rubric.md 的公式复算 credibility，
    # 自报值仅为初值，不一致则覆盖 + WARN + 计数（策略：纠正降级，不整单失败；
    # 只有下面 leads 全部失效才会走到已有的"线索为空"分支）
    warnings = []
    corrected_count = 0
    for ld in leads:
        if _recompute_credibility(ld, warnings):
            corrected_count += 1
    for w in warnings:
        print(f"WARN: {w}", file=sys.stderr)
    if corrected_count:
        meta["corrected_count"] = corrected_count

    expect = build(leads, meta, args.output)
    problems = verify(args.output, expect)
    if problems:
        print("VERIFY_FAIL: " + "; ".join(problems), file=sys.stderr); sys.exit(1)
    print(f"SUCCESS: {args.output}")
    print(f"  线索 {expect['total']} 条 | A {expect['tiers']['A']} / "
          f"B {expect['tiers']['B']} / C {expect['tiers']['C']}")
    if skipped:
        print(f"  跳过坏行 {skipped} 条（详见 stderr WARN，已记入报告元数据）")
    if corrected_count:
        print(f"  评分被代码纠正 {corrected_count} 条（详见 stderr WARN，已记入报告元数据）")


if __name__ == "__main__":
    main()
